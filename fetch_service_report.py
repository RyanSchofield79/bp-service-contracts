#!/usr/bin/env python3
"""
Business Pilot - Monthly Service Contracts Report
==================================================
Fetches all LIVE service contracts created up to the end of the previous
calendar month (i.e. contracts created in the current month are excluded).
Generates an Excel report sorted by contract number (descending).

Usage:
    BP_API_KEY=your-key python fetch_service_report.py

Output:
    Service_Report_<Month>_<Year>.xlsx
"""

import os
import json
import urllib.request
import urllib.error
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_KEY  = os.environ.get('BP_API_KEY', '')
BASE_URL = 'https://api-zapier.businesspilot.co.uk'

FETCH_FROM_DATE = datetime.utcnow().replace(year=datetime.utcnow().year - 5, hour=0, minute=0, second=0, microsecond=0)
_today = datetime.utcnow()
CUTOFF_DATE = _today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
_prev_month_last = (CUTOFF_DATE.replace(day=1) - __import__('datetime').timedelta(days=1))
REPORT_MONTH_LABEL = _prev_month_last.strftime('%B_%Y')
REPORT_MONTH_HUMAN = _prev_month_last.strftime('%B %Y')

COLUMNS = ['Contract Date','Contract Number','Customer','Current Owner','Contract Type','Product Type','Install Address 1','Install Postcode','Net Value','Install Start','Contract Status','Contract Status Date','Balance']

def api_post(endpoint, body):
    url = f'{BASE_URL}/api{endpoint}'
    payload = json.dumps(body).encode('utf-8')
    req = urllib.request.Request(url, data=payload, method='POST', headers={'X-Api-Key': API_KEY, 'Content-Type': 'application/json', 'Accept': 'application/json'})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode('utf-8'))

def parse_date(s):
    if not s: return None
    try:
        s = s.replace('Z', '+00:00') if s.endswith('Z') else s
        return datetime.fromisoformat(s).replace(tzinfo=None)
    except: return None

def fetch_live_contracts():
    fetch_from_iso = FETCH_FROM_DATE.isoformat() + 'Z'
    resp = api_post('/Contracts/find', {'dateAddedAfter': fetch_from_iso, 'page': 1, 'pageSize': 10000})
    result = resp[0] if isinstance(resp, list) else resp
    all_contracts = result.get('items', [])
    live = [c for c in all_contracts if not c.get('isComplete') and not c.get('isCancelled')]
    service = [c for c in live if 'service' in (c.get('currentPipeline') or '').lower() or 'service' in ((c.get('lead') or {}).get('leadType') or '').lower()]
    filtered = [c for c in service if (parse_date(c.get('dateAdded')) or datetime.min) < CUTOFF_DATE]
    print(f'  {len(filtered)} service contracts after date filter (excluded {len(all_contracts)-len(live)} completed/cancelled, {len(live)-len(service)} non-service, {len(service)-len(filtered)} this month)')
    return filtered

def fmt_date(val):
    if not val: return ''
    try:
        s = val.replace('Z', '+00:00') if val.endswith('Z') else val
        return datetime.fromisoformat(s).replace(tzinfo=None)
    except: return ''

def _first(*args):
    for v in args:
        if v: return v
    return ''

def map_contract(c):
    lead = c.get('lead') or {}
    contact = lead.get('contact') or {}
    customer = (contact.get('contactName') or contact.get('companyName') or '').strip()
    product_type = _first(lead.get('productType1'), lead.get('productType2'), lead.get('productType3'), lead.get('productInterest'), c.get('productType'), c.get('product'), c.get('productDescription'), c.get('jobType'))
    install_addr = _first(c.get('installAddressLine1'), c.get('installAddress1'), c.get('siteAddressLine1'), c.get('siteAddress1'), c.get('address1'), (c.get('installAddress') or {}).get('addressLine1'), lead.get('installAddressLine1'), lead.get('installAddress1'), lead.get('siteAddressLine1'), lead.get('siteAddress1'))
    install_postcode = _first(lead.get('installAddressPostcode'), c.get('installPostcode'), c.get('installPostCode'), c.get('postcode'), c.get('postCode'))
    install_start = fmt_date(_first(c.get('installStart'), c.get('estimatedInstallDate'), lead.get('installStartDate'), lead.get('installDate')))
    balance = float(c.get('balance') or c.get('balanceDue') or 0)
    return {'Contract Date': fmt_date(c.get('dateAdded')), 'Contract Number': str(c.get('contractNumber', '')), 'Customer': customer, 'Current Owner': c.get('currentOwnerName', ''), 'Contract Type': lead.get('leadType', ''), 'Product Type': product_type, 'Install Address 1': install_addr, 'Install Postcode': install_postcode, 'Net Value': float(c.get('confirmedNetSaleValue') or 0), 'Install Start': install_start, 'Contract Status': c.get('currentStatus', ''), 'Contract Status Date': fmt_date(c.get('currentStatusDate')), 'Balance': balance}

HEADER_FILL = PatternFill('solid', fgColor='C00000')
GRAND_FILL = PatternFill('solid', fgColor='C00000')
ALT_ROW_FILL = PatternFill('solid', fgColor='F2F2F2')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
YELLOW_FILL = PatternFill('solid', fgColor='FFFF00')
RED_FILL = PatternFill('solid', fgColor='FF0000')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
COL_WIDTHS = [16, 16, 30, 20, 20, 20, 30, 14, 14, 16, 28, 22, 14]
DATE_FMT = 'DD/MM/YYYY'
DATETIME_FMT = 'DD/MM/YYYY HH:MM'
CURRENCY_FMT = '£#,##0.00'
THREE_WEEKS_DAYS = 35
SIX_WEEKS_DAYS = 63

def style_cell(cell, fill=None, bold=False, color='000000', size=10, h_align='left', border=True, number_format=None):
    cell.font = Font(name='Arial', bold=bold, color=color, size=size)
    if fill: cell.fill = fill
    cell.alignment = Alignment(horizontal=h_align, vertical='center')
    if border: cell.border = THIN_BORDER
    if number_format: cell.number_format = number_format

def generate_excel(contracts, output_path):
    rows = sorted([map_contract(c) for c in contracts], key=lambda r: int(r['Contract Number']) if r['Contract Number'].isdigit() else 0, reverse=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Live Service Report'
    for i, width in enumerate(COL_WIDTHS, 1): ws.column_dimensions[get_column_letter(i)].width = width
    ws.append(COLUMNS)
    for cell in ws[1]: style_cell(cell, fill=HEADER_FILL, bold=True, color='FFFFFF', h_align='center', size=10)
    ws.freeze_panes = 'A2'
    data_start = 2
    now = datetime.now()
    for idx, row in enumerate(rows):
        ws.append([row['Contract Date'], row['Contract Number'], row['Customer'], row['Current Owner'], row['Contract Type'], row['Product Type'], row['Install Address 1'], row['Install Postcode'], row['Net Value'], row['Install Start'], row['Contract Status'], row['Contract Status Date'], row['Balance']])
        status_date = row['Contract Status Date']
        if isinstance(status_date, datetime) and status_date:
            age = (now - status_date).days
            row_fill = RED_FILL if age >= SIX_WEEKS_DAYS else (YELLOW_FILL if age >= THREE_WEEKS_DAYS else (ALT_ROW_FILL if idx % 2 == 1 else WHITE_FILL))
            text_color = 'FFFFFF' if age >= SIX_WEEKS_DAYS else '000000'
        else:
            row_fill = ALT_ROW_FILL if idx % 2 == 1 else WHITE_FILL
            text_color = '000000'
        current_row = data_start + idx
        for col, cell in enumerate(ws[current_row], 1):
            fmt = CURRENCY_FMT if col in (9, 13) else (DATETIME_FMT if col in (1, 12) else (DATE_FMT if col == 10 else None))
            style_cell(cell, fill=row_fill, color=text_color, number_format=fmt)
    data_end = data_start + len(rows) - 1
    grand_row = data_end + 2
    ws.cell(grand_row, 1, 'GRAND TOTAL')
    ws.cell(grand_row, 2, len(rows))
    ws.cell(grand_row, 9, f'=SUM(I{data_start}:I{data_end})')
    ws.cell(grand_row, 13, f'=SUM(M{data_start}:M{data_end})')
    for col in range(1, 14):
        cell = ws.cell(grand_row, col)
        style_cell(cell, fill=GRAND_FILL, bold=True, color='FFFFFF', number_format=CURRENCY_FMT if col in (9, 13) else None)
    wb.save(output_path)
    print(f'Report saved: {output_path}')

if __name__ == '__main__':
    if not API_KEY:
        print('ERROR: BP_API_KEY environment variable not set.')
        raise SystemExit(1)
    output_filename = f'Service_Report_{REPORT_MONTH_LABEL}.xlsx'
    contracts = fetch_live_contracts()
    if not contracts:
        print('No live service contracts found - nothing to report.')
        raise SystemExit(0)
    generate_excel(contracts, output_filename)
    print(f'\nReport covers: live service contracts created up to end of {REPORT_MONTH_HUMAN}')
    print(f'Total contracts in report: {len(contracts)}')
